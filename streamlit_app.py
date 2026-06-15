#!/usr/bin/env python3
"""
PDF Quote Extractor - Streamlit Web App
从PDF报价单中提取信息并导出Excel
"""

import streamlit as st
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import json
from pathlib import Path

# 页面配置
st.set_page_config(
    page_title="PDF 报价提取工具",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义样式
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 2px;
    }
    </style>
    """, unsafe_allow_html=True)

TARGET_COLUMNS = [
    'Numéro',
    'Libellé / Désignation',
    'Unité',
    'Quantité',
    'PU EXW excluding tax $',
    'Montant Total EXW excluding tax $'
]

class QuoteExtractor:
    def __init__(self, pdf_file):
        self.pdf_file = pdf_file
        self.tables = []
        self.text_content = ""
        self.extracted_data = []
    
    def extract(self):
        """主提取流程"""
        try:
            with pdfplumber.open(self.pdf_file) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    # 提取表格
                    page_tables = page.extract_tables()
                    if page_tables:
                        for table in page_tables:
                            self.tables.append({
                                'page': page_num,
                                'data': table
                            })
                    
                    # 提取文本
                    text = page.extract_text() or ""
                    self.text_content += f"\n--- Page {page_num} ---\n{text}"
            
            self._find_and_parse_quote_table()
            return True
        
        except Exception as e:
            st.error(f"❌ 读取PDF错误: {e}")
            return False
    
    def _find_and_parse_quote_table(self):
        """查找并解析报价表格"""
        if not self.tables:
            st.warning("⚠️ 未找到表格，尝试从文本提取...")
            return
        
        quote_table_info = self._identify_quote_table()
        
        if quote_table_info:
            self._parse_table_based_quote(quote_table_info)
        else:
            largest_table = max(self.tables, key=lambda x: len(x['data']))
            self._parse_table_based_quote(largest_table)
    
    def _identify_quote_table(self):
        """智能识别报价表格"""
        for table_info in self.tables:
            table = table_info['data']
            if len(table) < 2:
                continue
            
            header = ' '.join(str(cell).lower() for cell in table[0] if cell)
            
            keywords = ['numéro', 'désignation', 'description', 'unité', 
                       'quantité', 'qty', 'prix', 'pu', 'montant', 'total']
            matches = sum(1 for kw in keywords if kw in header)
            
            if matches >= 3:
                return table_info
        
        return None
    
    def _parse_table_based_quote(self, table_info):
        """从表格解析报价数据"""
        table = table_info['data']
        
        cleaned_table = self._clean_table(table)
        
        if len(cleaned_table) < 2:
            return
        
        header = cleaned_table[0]
        col_mapping = self._map_columns(header)
        
        for row in cleaned_table[1:]:
            if not any(cell.strip() for cell in row):
                continue
            
            extracted_row = self._map_row_to_target(row, col_mapping)
            if any(extracted_row):
                self.extracted_data.append(extracted_row)
    
    def _clean_table(self, table):
        """清理表格数据"""
        cleaned = []
        for row in table:
            if any(cell for cell in row if cell):
                cleaned_row = [
                    str(cell).strip() if cell else ""
                    for cell in row
                ]
                cleaned.append(cleaned_row)
        return cleaned
    
    def _map_columns(self, header_row):
        """智能列映射"""
        mapping = {}
        header_lower = [str(h).lower() for h in header_row]
        
        rules = {
            'Numéro': ['numéro', 'n°', 'num', 'reference', 'ref', 'article'],
            'Libellé / Désignation': ['libellé', 'désignation', 'description', 'item', 'product'],
            'Unité': ['unité', 'unit', 'u', 'um'],
            'Quantité': ['quantité', 'qty', 'qté', 'quantity', 'q'],
            'PU EXW excluding tax $': ['pu', 'prix', 'price', 'unit price', 'unitaire', 'pu ht'],
            'Montant Total EXW excluding tax $': ['montant', 'total', 'amount', 'montant total'],
        }
        
        for target_col, keywords in rules.items():
            for idx, header in enumerate(header_lower):
                if any(kw in header for kw in keywords):
                    mapping[idx] = target_col
                    break
        
        return mapping
    
    def _map_row_to_target(self, row, col_mapping):
        """将行数据映射到目标列"""
        result = [''] * len(TARGET_COLUMNS)
        
        for src_idx, target_col_name in col_mapping.items():
            if src_idx < len(row):
                target_idx = TARGET_COLUMNS.index(target_col_name)
                result[target_idx] = row[src_idx]
        
        return result
    
    def to_excel(self):
        """生成Excel字节流"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Devis"
        
        # 表头
        for col_idx, col_name in enumerate(TARGET_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 数据
        for row_idx, row_data in enumerate(self.extracted_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                if col_idx in [4, 5, 6]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    try:
                        if value and str(value).strip():
                            numeric = float(str(value).replace(',', '.').strip())
                            cell.value = numeric
                            cell.number_format = '#,##0.00'
                    except:
                        pass
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        column_widths = {'A': 12, 'B': 35, 'C': 12, 'D': 12, 'E': 18, 'F': 22}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        ws.freeze_panes = "A2"
        
        # 返回字节流
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def to_json(self):
        """生成JSON格式数据"""
        return {
            'total_rows': len(self.extracted_data),
            'columns': TARGET_COLUMNS,
            'data': [
                dict(zip(TARGET_COLUMNS, row))
                for row in self.extracted_data
            ]
        }
    
    def to_dataframe(self):
        """转换为pandas DataFrame"""
        return pd.DataFrame(self.extracted_data, columns=TARGET_COLUMNS)


def main():
    # 标题
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title("📄 PDF 报价提取工具")
        st.markdown("*从PDF报价单中自动提取信息，导出为Excel表格*")
    
    with col2:
        st.markdown("### ⚙️")
        st.info("v1.0", icon="ℹ️")
    
    st.divider()
    
    # 侧边栏配置
    with st.sidebar:
        st.header("📋 配置")
        
        show_advanced = st.checkbox("高级选项", value=False)
        
        if show_advanced:
            st.subheader("列配置")
            st.info("当前支持的列：\n" + "\n".join([f"• {col}" for col in TARGET_COLUMNS]))
        
        st.divider()
        st.markdown("### 📚 帮助")
        with st.expander("如何使用？", expanded=False):
            st.markdown("""
            1. **上传PDF** - 选择报价单PDF文件
            2. **自动识别** - 工具自动检测和提取表格
            3. **审查数据** - 查看提取结果，必要时调整
            4. **下载结果** - 导出为Excel或JSON
            """)
        
        with st.expander("支持的格式", expanded=False):
            st.markdown("""
            **输入：** PDF格式的报价单（需包含可提取的文本表格）
            
            **输出：** 
            - Excel (.xlsx) - 专业格式
            - JSON (.json) - 便于调整
            """)
        
        with st.expander("故障排除", expanded=False):
            st.markdown("""
            ❓ **找不到表格？**
            - 确保PDF包含文本内容（非扫描版）
            - 在高级选项中查看提取的文本
            
            ❓ **列识别不正确？**
            - 检查PDF报价单的列标题
            - 必要时手动调整JSON数据
            
            ❓ **数字未正确解析？**
            - 导出JSON后手动调整数字格式
            """)
    
    # 主要区域
    tab1, tab2, tab3 = st.tabs(["📤 上传 & 提取", "📊 审查数据", "💾 导出结果"])
    
    # 初始化session state
    if 'extractor' not in st.session_state:
        st.session_state.extractor = None
    if 'extracted' not in st.session_state:
        st.session_state.extracted = False
    
    # Tab 1: 上传和提取
    with tab1:
        st.header("第1步：上传PDF文件")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            uploaded_file = st.file_uploader(
                "选择PDF文件",
                type=['pdf'],
                help="上传报价单PDF（必须包含可提取的表格）"
            )
        
        with col2:
            st.markdown("### 或")
        
        if uploaded_file:
            st.success(f"✅ 已选择文件：{uploaded_file.name}")
            
            st.header("第2步：提取数据")
            
            if st.button("🚀 开始提取", type="primary", use_container_width=True):
                with st.spinner("正在读取PDF并提取数据..."):
                    extractor = QuoteExtractor(uploaded_file)
                    
                    if extractor.extract():
                        st.session_state.extractor = extractor
                        st.session_state.extracted = True
                        
                        st.success(f"✅ 提取完成！")
                        st.info(f"📊 提取了 **{len(extractor.extracted_data)} 行** 数据")
                        
                        # 显示预览
                        st.subheader("📋 数据预览（前10行）")
                        df_preview = extractor.to_dataframe().head(10)
                        st.dataframe(df_preview, use_container_width=True, hide_index=True)
                    else:
                        st.error("❌ 提取失败，请检查PDF文件")
    
    # Tab 2: 审查数据
    with tab2:
        if st.session_state.extracted and st.session_state.extractor:
            st.header("📊 数据审查")
            
            extractor = st.session_state.extractor
            df = extractor.to_dataframe()
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("总行数", len(df))
            with col2:
                st.metric("总列数", len(TARGET_COLUMNS))
            with col3:
                st.metric("完整度", f"{(df.notna().sum().sum() / (len(df) * len(TARGET_COLUMNS)) * 100):.1f}%")
            with col4:
                st.metric("空行数", df.isna().all(axis=1).sum())
            
            st.divider()
            
            # 完整数据表
            st.subheader("📑 完整数据")
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            # 数据统计
            st.subheader("📈 数据统计")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**每列的非空数据数**")
                col_stats = df.notna().sum()
                st.bar_chart(col_stats)
            
            with col2:
                st.markdown("**数据完整性**")
                completeness = (df.notna().sum() / len(df) * 100).round(1)
                st.bar_chart(completeness)
            
            # 数据质量检查
            st.subheader("🔍 数据质量检查")
            
            quality_checks = []
            
            # 检查关键列
            if df['Numéro'].isna().any():
                quality_checks.append("⚠️ Numéro 列有缺失值")
            
            if df['Libellé / Désignation'].isna().any():
                quality_checks.append("⚠️ 描述列有缺失值")
            
            # 检查数字列
            for col in ['Quantité', 'PU EXW excluding tax $', 'Montant Total EXW excluding tax $']:
                try:
                    df[col].astype(float)
                except:
                    quality_checks.append(f"⚠️ {col} 包含非数字值")
            
            if quality_checks:
                for check in quality_checks:
                    st.warning(check)
            else:
                st.success("✅ 所有数据看起来都很好！")
        
        else:
            st.info('💡 请先在"上传 & 提取"标签页上传并提取数据')
    
    # Tab 3: 导出结果
    with tab3:
        if st.session_state.extracted and st.session_state.extractor:
            st.header("💾 导出数据")
            
            extractor = st.session_state.extractor
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("📊 导出为Excel")
                
                excel_data = extractor.to_excel()
                
                st.download_button(
                    label="⬇️ 下载 Excel (.xlsx)",
                    data=excel_data,
                    file_name="quote_extract.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                
                st.markdown("""
                **Excel文件包含：**
                - ✅ 蓝色表头，格式化专业
                - ✅ 冻结表头行
                - ✅ 自动列宽
                - ✅ 数字格式 (#,##0.00)
                """)
            
            with col2:
                st.subheader("📄 导出为JSON")
                
                json_data = extractor.to_json()
                json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
                
                st.download_button(
                    label="⬇️ 下载 JSON (.json)",
                    data=json_str,
                    file_name="quote_extract.json",
                    mime="application/json",
                    use_container_width=True
                )
                
                st.markdown("""
                **JSON文件包含：**
                - ✅ 结构化数据
                - ✅ 便于编程处理
                - ✅ 支持进一步分析
                """)
            
            st.divider()
            
            # 预览JSON
            with st.expander("👀 预览JSON", expanded=False):
                st.json(json_data)
            
            # CSV选项
            st.subheader("📋 或导出为CSV")
            
            csv_data = extractor.to_dataframe().to_csv(index=False, encoding='utf-8-sig')
            
            st.download_button(
                label="⬇️ 下载 CSV (.csv)",
                data=csv_data,
                file_name="quote_extract.csv",
                mime="text/csv",
                use_container_width=True
            )
        
        else:
            st.info('💡 请先在"上传 & 提取"标签页上传并提取数据')
    
    # 页脚
    st.divider()
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**📝 版本：** 1.0")
    
    with col2:
        st.markdown("**🔧 技术栈：** Streamlit + pdfplumber + openpyxl")
    
    with col3:
        st.markdown("**👤 作者：** Zélie")


if __name__ == "__main__":
    main()
